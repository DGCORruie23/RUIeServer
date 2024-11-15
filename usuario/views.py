from django.shortcuts import render, get_object_or_404, redirect
from django.db.models import Max, Count
from django.views.decorators.csrf import csrf_exempt
from rest_framework.parsers import JSONParser
from django.http import JsonResponse, HttpResponseRedirect, HttpResponse
from django.contrib.auth.hashers import make_password, check_password
from .serializers import UserGetSerializer, UserGetSerializerC, PaisesGetSerializer, EstadoFuerzaGetSerializer, FrasesGetSerializer, ListRescatePuntoSerializer, RescatePuntoSerializer
from .serializers import MunicipiosGetSerializer, PuntosInterGetSerializer, ConteoRapidoSerializer, MsgUpdateGetSerializer, ConteoDisuadidosSerializer
from .models import Usuario, Paises, EstadoFuerza, Frases, Municipios, PuntosInternacion, RescatePunto, ConteoRapidoPunto, MsgUpdate
from .forms import CargarArchivoForm, ExcelForm, ExcelFormOr, ExcelFormOrs 
import openpyxl as opxl
from openpyxl.writer.excel import save_virtual_workbook
from datetime import *
from django.contrib.auth.decorators import login_required

# Create your views here.

def detail(request, question_id):
  question = Question.objects.get(pk=question_id)
  
@csrf_exempt
def login_user(request):

    if request.method == 'POST':
        datos_recibidos = JSONParser().parse(request)
        datos_serializados = UserGetSerializer(data=datos_recibidos)
        # print(datos_serializados)
        if datos_serializados.is_valid():
            datos = datos_serializados.data
            usuario_info = datos.get('nickname')
            pass_info = datos.get('password')

            # print(f" nickname: {usuario_info}, pass: {pass_info}")

            if(Usuario.objects.filter(nickname = usuario_info).exists()):

                datos_usuario = Usuario.objects.get(nickname = usuario_info)

                if(check_password(pass_info, datos_usuario.password)):
                    nuevo_serializer = {'nickname': datos_usuario.nickname,
                                        'nombre': datos_usuario.nombre,
                                        'apellido': datos_usuario.apellido,
                                        'password': 'ok',
                                        'estado': datos_usuario.estado,
                                        'tipo': datos_usuario.tipo}
                    
                    # print(f" ${datos_usuario.nickname} ${datos_usuario.nombre} ${datos_usuario.apellido} ${datos_usuario.estado}")
                    serializer_datos = UserGetSerializerC(nuevo_serializer)
                    return JsonResponse(serializer_datos.data, status = 200)

                else:
                    nuevo_serializer = {'nickname': 'error',
                                        'nombre': '',
                                        'apellido': '',
                                         'password': 'error',
                                         'estado': '',
                                         'tipo': ''}
                    serializer_datos = UserGetSerializerC(nuevo_serializer)
                    return JsonResponse(serializer_datos.data, status = 200)

            else:

                nuevo_serializer = {'nickname': 'error',
                                        'nombre': '',
                                        'apellido': '',
                                         'password': 'error',
                                         'estado': '',
                                         'tipo': ''}
                serializer_datos = UserGetSerializerC(nuevo_serializer)
                return JsonResponse(serializer_datos.data, status = 200)
        else:
            nuevo_serializer = {'nickname': 'error',
                                        'nombre': '',
                                        'apellido': '',
                                         'password': 'error',
                                         'estado': '',
                                         'tipo': ''}
            serializer_datos = UserGetSerializerC(nuevo_serializer)
            return JsonResponse(serializer_datos.data, status = 200)

@csrf_exempt
def cargarPais(request):
    if(request.method == "POST"):
        form = CargarArchivoForm(request.POST, request.FILES)
        if(form.is_valid()):
            # Nota: "archivo" este campo se llama como se llama en el form  
            excel_file = request.FILES["archivo"]
            nombreA = str(excel_file.name)
            extensionA = (nombreA.split(".")[-1]).lower()
            # print((nombreA.split(".")[-1]).lower())
            if( extensionA == "xlsx" 
               or extensionA == ".xlsm" 
               or extensionA == ".xlsb" 
               or extensionA == ".xltx" 
               or extensionA == ".xltm" 
               or extensionA == ".xls"):
                dataWB = opxl.load_workbook(excel_file, data_only=True)

                data = dataWB.worksheets[0]

                # print(data.cell(4,2).value)

                # Se leen los datos del excel
                paises = []
                i = 1
                while not(i == 0):
                    if(data.cell(i+3,2).value == None):
                        i = 0
                    else:
                        paises.append([data.cell(i+3,2).value, data.cell(i+3,3).value])
                        i+=1
                #-----------------------------------

                #Eliminar todos los registros de Paises
                Paises.objects.all().delete()
                #------------------------

                # Se agrega a la DB los paises y se crea un string de ellos
                paisesStr = "{"    
                for num, nac in enumerate(paises):
                    paisesStr += f'\n\t"{nac}",' 
                    Paises.objects.create(idPais = num + 1, nombre_pais = nac[0], iso3 = nac[1])
                    # print(nac)
                paisesStr += "\n}" 
                #----------------------------

                # pasar string de las nacionalidades 
                # print(paisesStr) 
                #------------------------
            return redirect("dashboard")
    else:
        form = CargarArchivoForm()
    return render(request, "cargarExcel/cargarArchivoPaises.html", {"form" : form})




            

@csrf_exempt
def cargarPuntoI(request):
    if(request.method == "POST"):
        form = CargarArchivoForm(request.POST, request.FILES)
        if(form.is_valid()):
            # Nota: "archivo" este campo se llama como se llama en el form  
            excel_file = request.FILES["archivo"]
            nombreA = str(excel_file.name)
            extensionA = (nombreA.split(".")[-1]).lower()
            # print((nombreA.split(".")[-1]).lower())
            if( extensionA == "xlsx" 
               or extensionA == ".xlsm" 
               or extensionA == ".xlsb" 
               or extensionA == ".xltx" 
               or extensionA == ".xltm" 
               or extensionA == ".xls"):
                dataWB = opxl.load_workbook(excel_file, data_only=True)

                data = dataWB.worksheets[0]

                # print(data.cell(4,2).value)

                # Se leen los datos del excel
                puntoI = []
                i = 1
                while not(i == 0):
                    if(data.cell(i,1).value == None):
                        i = 0
                    else:
                        puntoI.append([data.cell(i,1).value, data.cell(i,2).value, data.cell(i,3).value])
                        i+=1
                #-----------------------------------

                #Eliminar todos los registros de Paises
                PuntosInternacion.objects.all().delete()
                #------------------------

                # Se agrega a la DB los paises y se crea un string de ellos
                puntoIStr = "{"    
                for num, punto in enumerate(puntoI):
                    puntoIStr += f'\n\t"{punto}",' 
                    PuntosInternacion.objects.create(idPuntoInter = num + 1, estadoPunto = punto[0], nombrePunto = punto[1], tipoPunto = punto[2], )
                    # print(nac)
                puntoIStr += "\n}" 
                #----------------------------

                # pasar string de las nacionalidades 
                # print(puntoIStr) 
                #------------------------
            return redirect("dashboard")
    else:
        form = CargarArchivoForm()
    return render(request, "cargarExcel/cargarArchivoPuntosInter.html", {"form" : form})



@csrf_exempt
def cargarEdoFuerza(request):
    if(request.method == "POST"):
        form = CargarArchivoForm(request.POST, request.FILES)
        if(form.is_valid()):
            # Nota: "archivo" este campo se llama como se llama en el form  
            excel_file = request.FILES["archivo"]
            nombreA = str(excel_file.name)
            extensionA = (nombreA.split(".")[-1]).lower()
            # print((nombreA.split(".")[-1]).lower())
            if( extensionA == "xlsx" 
               or extensionA == ".xlsm" 
               or extensionA == ".xlsb" 
               or extensionA == ".xltx" 
               or extensionA == ".xltm" 
               or extensionA == ".xls"):
                dataWB = opxl.load_workbook(excel_file, data_only=True)

                data = dataWB.worksheets[0]

                # print(data.cell(4,2).value)

                # Se leen los datos del excel
                edoFuerza = []
                auxL = []
            
                i = 1
                while not(i == 0):

                    auxL.clear()

                    if(data.cell(i+2,1).value == None):
                        i = 0
                        # print(edoFuerza)
                    else:
                        for x in range(1, 14, 1):
                            
                            ax = data.cell(i+2, x).value
                            
                            if(x == 13):
                                if(ax == "CECO RIO BRAVO"):
                                    auxL.append(1)
                                elif(ax == "CECO SUCHIATE"):
                                    auxL.append(2)
                                else:
                                    auxL.append(3)
                            else:
                                if(ax == None or ax == "N/A"):
                                    auxL.append("0")
                                else:
                                    auxL.append(ax)

                        edoFuerza.append(auxL[:])
                        i+=1
                #-----------------------------------

                

                #Verificar si la lista esta vacia para no guardar información
                if(bool(edoFuerza)):

                    # print(edoFuerza)

                    #Eliminar todos los registros de Paises
                    EstadoFuerza.objects.all().delete()
                    #------------------------

                    # Se agrega a la DB los paises y se crea un string de ellos
                    edoFuerzaStr = "{"    
                    for num, fuerzaP in enumerate(edoFuerza):
                        
                        edoFuerzaStr += "\n\t"
                        
                        edoFuerzaStr += f"{fuerzaP}"
                        
                        textoC, lat, lon = convertirU(fuerzaP[5])

                        EstadoFuerza.objects.create(
                            idEdoFuerza = num + 1 ,
                            oficinaR = str(fuerzaP[0]),
                            numPunto = int(str(fuerzaP[1])),
                            nomPuntoRevision = str(fuerzaP[2]),
                            tipoP = str(fuerzaP[3]),
                            ubicacion = str(fuerzaP[4]),
                            coordenadasTexto = textoC,
                            latitud = float(lat),
                            longitud = float(lon),
                            personalINM = int(str(fuerzaP[6])),
                            personalSEDENA = int(str(fuerzaP[7])),
                            personalMarina = int(str(fuerzaP[8])),
                            personalGuardiaN = int(str(fuerzaP[9])),
                            personalOTROS = int(str(fuerzaP[10])),
                            vehiculos = int(str(fuerzaP[11])),
                            seccion = int(str(fuerzaP[12])),
                            )
                        # print(nac)
                    edoFuerzaStr += "\n}" 
                    #----------------------------

                    # pasar string de las nacionalidades 
                    # print(edoFuerzaStr) 
                    #------------------------
            # return HttpResponseRedirect("/info/cargarFuerza")
            return redirect("pagina_pruebas_edoFuerza")
    else:
        form = CargarArchivoForm()
    return render(request, "cargarExcel/cargarArchivoFuerza.html", {"form" : form})





@csrf_exempt
def cargarMunicipios(request):
    if(request.method == "POST"):
        form = CargarArchivoForm(request.POST, request.FILES)
        if(form.is_valid()):
            # Nota: "archivo" este campo se llama como se llama en el form  
            excel_file = request.FILES["archivo"]
            nombreA = str(excel_file.name)
            extensionA = (nombreA.split(".")[-1]).lower()
            # print((nombreA.split(".")[-1]).lower())
            if( extensionA == "xlsx" 
               or extensionA == ".xlsm" 
               or extensionA == ".xlsb" 
               or extensionA == ".xltx" 
               or extensionA == ".xltm" 
               or extensionA == ".xls"):
                dataWB = opxl.load_workbook(excel_file, data_only=True)

                data = dataWB.worksheets[0]

                # print(data.cell(4,2).value)

                # Se leen los datos del excel
                municipio = []
                auxL = []
            
                i = 1
                while not(i == 0):

                    auxL.clear()

                    if(data.cell( i+4, 4).value == None):
                        i = 0
                        # print(edoFuerza)
                    else:
                        edo = data.cell(i + 3, 4).value
                        edoAbrev = data.cell(i + 4, 4).value
                        nomMun = data.cell(i + 4, 6).value

                        abrevMayus = ((edoAbrev.upper()).replace(".","")).replace(" ","")

                        if(abrevMayus == "AGS"):
                            edo = "AGUASCALIENTES"
                        elif(abrevMayus == "BC"):
                            edo = "BAJA CALIFORNIA"
                        elif(abrevMayus == "BCS"):
                            edo = "BAJA CALIFORNIA SUR"
                        elif(abrevMayus == "CAMP"):
                            edo = "CAMPECHE"
                        elif(abrevMayus == "CDMX"):
                            edo = "CDMX"
                        elif(abrevMayus == "CHIH"):
                            edo = "CHIHUAHUA"
                        elif(abrevMayus == "CHIS"):
                            edo = "CHIAPAS"
                        elif(abrevMayus == "COAH"):
                            edo = "COAHUILA"
                        elif(abrevMayus == "COL"):
                            edo = "COLIMA"
                        elif(abrevMayus == "DGO"):
                            edo = "DURANGO"
                        elif(abrevMayus == "GRO"):
                            edo = "GUERRERO"
                        elif(abrevMayus == "GTO"):
                            edo = "GUANAJUATO"
                        elif(abrevMayus == "HGO"):
                            edo = "HIDALGO"
                        elif(abrevMayus == "JAL"):
                            edo = "JALISCO"
                        elif(abrevMayus == "MEX"):
                            edo = "EDOMEX"
                        elif(abrevMayus == "MICH"):
                            edo = "MICHOACÁN"
                        elif(abrevMayus == "MOR"):
                            edo = "MORELOS"
                        elif(abrevMayus == "NAY"):
                            edo = "NAYARIT"
                        elif(abrevMayus == "NL"):
                            edo = "NUEVO LEÓN"
                        elif(abrevMayus == "OAX"):
                            edo = "OAXACA"
                        elif(abrevMayus == "PUE"):
                            edo = "PUEBLA"
                        elif(abrevMayus == "QROO"):
                            edo = "QUINTANA ROO"
                        elif(abrevMayus == "QRO"):
                            edo = "QUERÉTARO"
                        elif(abrevMayus == "SIN"):
                            edo = "SINALOA"
                        elif(abrevMayus == "SLP"):
                            edo = "SAN LUIS POTOSÍ"
                        elif(abrevMayus == "SON"):
                            edo = "SONORA"
                        elif(abrevMayus == "TAB"):
                            edo = "TABASCO"
                        elif(abrevMayus == "TAMPS"):
                            edo = "TAMAULIPAS"
                        elif(abrevMayus == "TLAX"):
                            edo = "TLAXCALA"
                        elif(abrevMayus == "VER"):
                            edo = "VERACRUZ"
                        elif(abrevMayus == "YUC"):
                            edo = "YUCATÁN"
                        elif(abrevMayus == "ZAC"):
                            edo = "ZACATECAS"
                        else:
                            edo = ""
                        
                        municipio.append([edo, abrevMayus, nomMun])

                        i+=1
                #-----------------------------------

                

                #Verificar si la lista esta vacia para no guardar información
                if(bool(municipio)):

                    # print(edoFuerza)

                    #Eliminar todos los registros de Paises
                    Municipios.objects.all().delete()
                    #------------------------

                    # Se agrega a la DB los paises y se crea un string de ellos
                    municipioStr = "{"    
                    for num, fuerzaP in enumerate(municipio):
                        
                        municipioStr += "\n\t"
                        
                        municipioStr += f"{fuerzaP}"

                        Municipios.objects.create(
                            idMunicipio = num + 1 ,
                            estado = str(fuerzaP[0]),
                            estadoAbr = str(fuerzaP[1]),
                            nomMunicipio = str(fuerzaP[2]),
                            )
                        # print(nac)
                    municipioStr += "\n}" 

                    # print("\n\nmunicipios")
                    # datosMun = Municipios.objects.values('estadoAbr').order_by('estadoAbr').distinct() 
                    # for x in datosMun:
                    #     print(x)

                    # print("\n\nOrs")
                    # datosOr = EstadoFuerza.objects.values('oficinaR').order_by('oficinaR').distinct() 
                    # for x in datosOr:
                    #     print(x)
                    #----------------------------

                    # pasar string de las nacionalidades 
                    # print(edoFuerzaStr) 
                    #------------------------
            return redirect("dashboard")
    else:
        form = CargarArchivoForm()
    return render(request, "cargarExcel/cargarArchivoMunicipios.html", {"form" : form})



@csrf_exempt
def cargaMasivaUser(request):
    if(request.method == "POST"):
        form = CargarArchivoForm(request.POST, request.FILES)
        if(form.is_valid()):
            # Nota: "archivo" este campo se llama como se llama en el form  
            excel_file = request.FILES["archivo"]
            nombreA = str(excel_file.name)
            extensionA = (nombreA.split(".")[-1]).lower()
            # print((nombreA.split(".")[-1]).lower())
            if( extensionA == "xlsx" 
               or extensionA == ".xlsm" 
               or extensionA == ".xlsb" 
               or extensionA == ".xltx" 
               or extensionA == ".xltm" 
               or extensionA == ".xls"):
                dataWB = opxl.load_workbook(excel_file, data_only=True)

                data = dataWB.worksheets[0]

                # print(data.cell(4,2).value)

                # Se leen los datos del excel
                municipio = []
                auxL = []
            
                i = 1
                while not(i == 0):

                    auxL.clear()

                    if(data.cell( i+4, 4).value == None):
                        i = 0
                        # print(edoFuerza)
                    else:
                        nickname = data.cell(i + 4, 2).value
                        password = data.cell(i + 4, 3).value
                        nombres = data.cell(i + 4, 4).value
                        apellidos = data.cell(i + 4, 5).value
                        estado = data.cell(i + 4, 6).value
                        tipo = data.cell(i + 4, 7).value

                        passUpdate = password                        

                        if (Usuario.objects.filter(nickname = nickname).exists()):
                            passUpdate = make_password(password)
                            Usuario.objects.filter(nickname = nickname).update(nombre=nombres, apellido=apellidos, password=passUpdate, estado=estado)
                        else:
                            Usuario.objects.create(
                                nickname=nickname, 
                                nombre=nombres, 
                                apellido=apellidos, 
                                password=passUpdate,
                                estado=estado,
                                tipo=tipo,
                            )
                        i = i + 1

                #-----------------------------------
            return redirect("pagina_pruebas_usuarios")
    else:
        form = CargarArchivoForm()
    return render(request, "cargarExcel/cargarUsuarios.html", {"form" : form})



@csrf_exempt
def insert_rescates(request):

    if request.method == 'POST':
        datos_recibidos = JSONParser().parse(request)
        datos_serializados = RescatePuntoSerializer(data=datos_recibidos, many=True)
        if datos_serializados.is_valid():
            datos_serializados.save()

            # print(datos)
            dataR = {
                'guardado' : "ok"
            }
            return JsonResponse(dataR, status = 200)
            
        else:
            print(datos_serializados.data)
            datos_serializados.is_valid
            print(datos_serializados.errors)
            dataR = {
                'guardado' : "error"
            }
            return JsonResponse( dataR, status = 200)

@csrf_exempt
def insert_conteo(request):

    if request.method == 'POST':
        datos_recibidos = JSONParser().parse(request)
        datos_serializados = ConteoRapidoSerializer(data=datos_recibidos, many=True)
        if datos_serializados.is_valid():
            datos_serializados.save()
            # print(datos)
            dataR = {
                'guardado' : "ok"
            }
            return JsonResponse(dataR, status = 200)
            
        else:
            # print(datos_serializados.data)
            # datos_serializados.is_valid
            # print(datos_serializados.errors)
            dataR = {
                'guardado' : "error"
            }
            return JsonResponse( dataR, status = 200)
        
@csrf_exempt
def insert_disuadidos(request):
    if request.method == 'POST':
        datos_recibidos = JSONParser().parse(request)
        datos_serializados = ConteoDisuadidosSerializer(data=datos_recibidos, many=True)
        if datos_serializados.is_valid():
            datos_serializados.save()
            # print(datos)
            dataR = {
                'guardado' : "ok"
            }
            return JsonResponse(dataR, status = 200)
            
        else:
            # print(datos_serializados.data)
            # datos_serializados.is_valid
            # print(datos_serializados.errors)
            dataR = {
                'guardado' : "error"
            }
            return JsonResponse(dataR, status = 400)

@csrf_exempt
def msgUpdateUrl(request):
    if request.method == 'GET':
        snippets = MsgUpdate.objects.last()
        serializer = MsgUpdateGetSerializer(snippets, many=False)
        return JsonResponse(serializer.data, safe=False)

@csrf_exempt
def infoPaises(request):
    if request.method == 'GET':
        snippets = Paises.objects.all()
        serializer = PaisesGetSerializer(snippets, many=True)
        return JsonResponse(serializer.data, safe=False)

@csrf_exempt
def infoMunicipios(request):
    if request.method == 'GET':
        snippets = Municipios.objects.all()
        serializer = MunicipiosGetSerializer(snippets, many=True)
        return JsonResponse(serializer.data, safe=False)
    
@csrf_exempt
def infoPuntosInterna(request):
    if request.method == 'GET':
        snippets = PuntosInternacion.objects.all()
        serializer = PuntosInterGetSerializer(snippets, many=True)
        return JsonResponse(serializer.data, safe=False)
    
@csrf_exempt
def infoEstadoFuerza(request):
    if request.method == 'GET':
        snippets = EstadoFuerza.objects.all()
        serializer = EstadoFuerzaGetSerializer(snippets, many=True)
        return JsonResponse(serializer.data, safe=False)
    
@csrf_exempt
def infoFrases(request):
    if request.method == 'GET':
        snippets = Frases.objects.all()
        serializer = FrasesGetSerializer(snippets, many=True)
        return JsonResponse(serializer.data, safe=False)

@csrf_exempt 
def convertirU(coordenadas):
    strCoor = str(coordenadas)
    lat = float("0.0")
    lon = float("0.0")
    return strCoor, lat, lon




@csrf_exempt
def generarExcelNombres(request):
    if(request.method == "POST"):
        form = ExcelForm(request.POST)
        if(form.is_valid()):
            # print(request.POST)
            # Nota: "archivo" este campo se llama como se llama en el form  
            dia = request.POST["fechaDescarga_day"]
            mes = request.POST["fechaDescarga_month"]
            year = request.POST["fechaDescarga_year"]

            fechaR = datetime.strptime(f"{dia}/{mes}/{year}", "%d/%m/%Y").strftime('%d-%m-%y')

            valores = RescatePunto.objects.filter(fecha= fechaR)
            
            workbook = opxl.load_workbook('tmp/dup.xlsm', read_only=False, keep_vba=True)
            worksheet = workbook.active
            
            for valor in valores:
                worksheet.append([valor.oficinaRepre, 
                                 valor.fecha,
                                 valor.hora,
                                 valor.nombreAgente.upper(),
                                 "1" if valor.aeropuerto else "",
                                 "1" if valor.carretero else "",
                                 valor.tipoVehic.upper(),
                                 valor.lineaAutobus.upper(),
                                 valor.numeroEcono.upper(),
                                 valor.placas.upper(),
                                 "1" if valor.vehiculoAseg else "",
                                 "1" if valor.casaSeguridad else "",
                                 "1" if valor.centralAutobus else "",
                                 "1" if valor.ferrocarril else "",
                                 valor.empresa,
                                 "1" if valor.hotel else "",
                                 valor.nombreHotel,
                                 "1" if valor.puestosADispo else "",
                                 "1" if valor.juezCalif else "",
                                 "1" if valor.reclusorio else "",
                                 "1" if valor.policiaFede else "",
                                 "1" if valor.dif else "",
                                 "1" if valor.policiaEsta else "",
                                 "1" if valor.policiaMuni else "",
                                 "1" if valor.guardiaNaci else "",
                                 "1" if valor.fiscalia else "",
                                 "1" if valor.otrasAuto else "",
                                 "1" if valor.voluntarios else "",
                                 "1" if valor.otro else "",
                                 "1" if valor.presuntosDelincuentes else "",
                                 valor.numPresuntosDelincuentes if valor.numPresuntosDelincuentes != 0 else "",
                                 valor.municipio.upper(),
                                 valor.puntoEstra.upper(),
                                 valor.nacionalidad.upper(),
                                 valor.iso3,
                                 valor.nombre.upper(),
                                 valor.apellidos.upper(),
                                 valor.noIdentidad,
                                 valor.parentesco.upper(),
                                 valor.fechaNacimiento,
                                 "Hombre" if valor.sexo else "Mujer",
                                 "1" if valor.embarazo else "",
                                 valor.numFamilia if valor.numFamilia != 0 else "",
                                 valor.edad,
                                 ])
            
            worksheet.append(['Rescates Totales: ', valores.count()])

            response = HttpResponse(content = save_virtual_workbook(workbook), content_type='application/vnd.ms-excel.sheet.macroEnabled.12')
            response['Content-Disposition'] = 'attachment; filename={fecha}.xlsm'.format(fecha = fechaR)

            return response
    else:
        form = ExcelForm()
    return render(request, "generarExcel/generarArchivoExcelNombres.html", {"form" : form})


@csrf_exempt
def generarExcelConteo(request):
    if(request.method == "POST"):
        form = ExcelForm(request.POST)
        print("entro")
        if(form.is_valid()):
            # print(request.POST)
            # Nota: "archivo" este campo se llama como se llama en el form  
            dia = request.POST["fechaDescarga_day"]
            mes = request.POST["fechaDescarga_month"]
            year = request.POST["fechaDescarga_year"]

            fechaR = datetime.strptime(f"{dia}/{mes}/{year}", "%d/%m/%Y").strftime('%d-%m-%y')

            valores = ConteoRapidoPunto.objects.filter(fecha= fechaR)
            
            workbook = opxl.load_workbook('tmp/dup.xlsm', read_only=False, keep_vba=True)
            worksheet = workbook.active

            for valor in valores:
                masivoD = (valor.AS_hombres + valor.AS_mujeresNoEmb +valor.AS_mujeresEmb +
                           valor.AA_hombres + valor.AA_mujeresNoEmb + valor.AA_mujeresEmb +
                           valor.NNA_A_hombres + valor.NNA_A_mujeresNoEmb + valor.NNA_A_mujeresEmb +
                           valor.NNA_S_hombres + valor.NNA_S_mujeresNoEmb + valor.NNA_S_mujeresEmb
                           )
                worksheet.append([valor.oficinaRepre, 
                                 valor.fecha,
                                 valor.hora,
                                 valor.nombreAgente.upper(),
                                 "1" if valor.aeropuerto else "",
                                 "1" if valor.carretero else "",
                                 valor.tipoVehic.upper(),
                                 valor.lineaAutobus.upper(),
                                 valor.numeroEcono.upper(),
                                 valor.placas.upper(),
                                 "1" if valor.vehiculoAseg else "",
                                 "1" if valor.casaSeguridad else "",
                                 "1" if valor.centralAutobus else "",
                                 "1" if valor.ferrocarril else "",
                                 valor.empresa,
                                 "1" if valor.hotel else "",
                                 valor.nombreHotel,
                                 "1" if valor.puestosADispo else "",
                                 "1" if valor.juezCalif else "",
                                 "1" if valor.reclusorio else "",
                                 "1" if valor.policiaFede else "",
                                 "1" if valor.dif else "",
                                 "1" if valor.policiaEsta else "",
                                 "1" if valor.policiaMuni else "",
                                 "1" if valor.guardiaNaci else "",
                                 "1" if valor.fiscalia else "",
                                 "1" if valor.otrasAuto else "",
                                 "1" if valor.voluntarios else "",
                                 "1" if valor.otro else "",
                                 "1" if valor.presuntosDelincuentes else "",
                                 valor.numPresuntosDelincuentes if valor.numPresuntosDelincuentes != 0 else "",
                                 valor.municipio.upper(),
                                 valor.puntoEstra.upper(),
                                 valor.nacionalidad.upper(),
                                 valor.iso3,

                                 valor.AS_hombres,
                                 valor.AS_mujeresNoEmb,
                                 valor.AS_mujeresEmb,

                                 valor.nucleosFamiliares,
                                 valor.AA_hombres,
                                 valor.AA_mujeresNoEmb,
                                 valor.AA_mujeresEmb,
                                 valor.NNA_A_hombres,
                                 valor.NNA_A_mujeresNoEmb,
                                 valor.NNA_A_mujeresEmb,

                                 valor.NNA_S_hombres,
                                 valor.NNA_S_mujeresNoEmb,
                                 valor.NNA_S_mujeresEmb,

                                 "1" if masivoD >= 40 else "",
                                 ])
            
            response = HttpResponse(content = save_virtual_workbook(workbook), content_type='application/vnd.ms-excel.sheet.macroEnabled.12')
            response['Content-Disposition'] = 'attachment; filename=rapido {fecha}.xlsm'.format(fecha = fechaR)
            return response
    else:
        form = ExcelForm()
    return render(request, "generarExcel/generarArchivoExcelConteo.html", {"form" : form})


def servirApps(request):
    return render(request, "descargas/descargar_Apps.html", {

    })

@csrf_exempt
def downloadAPK(request):
    if request.method == 'GET':
        appAndroid = open('tmp/ruie.apk', 'rb')

        response = HttpResponse(appAndroid, content_type="application/vnd.android.package-archive")
        response["Content-disposition"] = "attachment; filename=ruie.apk"
        return response
    
@csrf_exempt
def pagDuplicados(request):
    if request.method == 'GET':
        return render(request, "descargas/descargar_duplicados.html", {})
        
@csrf_exempt
def downloadDuplicados(request):
    if request.method == 'GET':

        workbook = opxl.load_workbook('tmp/rec.xlsm', read_only=False, keep_vba=True)
        worksheet = workbook.active

        fechaR = datetime.today().strftime('%d-%m-%y')
        
        # duplicados = RescatePunto.objects.all()

        register3 = RescatePunto.objects.values("iso3", "nombre", "apellidos", "fechaNacimiento").annotate(records=Count("*")).filter(records__gt=1)
        duplicados = []
        for reg in register3:
            for registro in RescatePunto.objects.filter(iso3=reg['iso3'], nombre=reg['nombre'], apellidos = reg['apellidos'], fechaNacimiento=reg['fechaNacimiento']):
                duplicados.append(registro)
                
        for valor in duplicados:
            puntoR = ""
            if valor.aeropuerto:
                puntoR = 'aeropuerto'
            elif valor.carretero:
                puntoR = 'carretero'
            elif valor.centralAutobus:
                puntoR = 'central de autobus'
            elif valor.casaSeguridad:
                puntoR = 'casa de seguridad'
            elif valor.ferrocarril:
                puntoR = 'ferrocarril'
            elif valor.hotel:
                puntoR = 'hotel'
            elif valor.puestosADispo:
                puntoR = 'puestos a disposicion'
            elif valor.voluntarios:
                puntoR = 'voluntarios'
            else: 
                puntoR = ''
            worksheet.append([valor.oficinaRepre, 
                              valor.fecha,
                              valor.hora,
                              valor.nombreAgente.upper(),
                              puntoR.upper(),
                              valor.puntoEstra.upper(),
                              valor.nacionalidad.upper(),
                              valor.iso3,
                              valor.nombre.upper(),
                              valor.apellidos.upper(),
                              valor.parentesco.upper(),
                              valor.fechaNacimiento,
                              valor.edad,
                              "Hombre" if valor.sexo else "Mujer",
                              "1" if valor.embarazo else "",
                              valor.numFamilia if valor.numFamilia != 0 else "",
                              ])
            
        worksheet['F1'] = 'Total Registros'
        worksheet['G1'] = str(RescatePunto.objects.count())

        response = HttpResponse(content = save_virtual_workbook(workbook), content_type='application/vnd.ms-excel.sheet.macroEnabled.12')
        response['Content-Disposition'] = 'attachment; filename=Reincidentes_a_{fecha}.xlsm'.format(fecha = fechaR)

        return response


@csrf_exempt
def generarExcelTab(request):
    if(request.method == "POST"):
        form = ExcelFormOr(request.POST)
        if(form.is_valid()):
            dia = request.POST["fechaDescarga_day"]
            mes = request.POST["fechaDescarga_month"]
            year = request.POST["fechaDescarga_year"]

            fechaR = datetime.strptime(f"{dia}/{mes}/{year}", "%d/%m/%Y").strftime('%d-%m-%y')

            valores = valores = RescatePunto.objects.filter(fecha= fechaR).filter( oficinaRepre="TABASCO")
            
            workbook = opxl.load_workbook('tmp/dup.xlsm', read_only=False, keep_vba=True)
            worksheet = workbook.active
            
            for valor in valores:
                worksheet.append([valor.oficinaRepre, 
                                 valor.fecha,
                                 valor.hora,
                                 valor.nombreAgente.upper(),
                                 "1" if valor.aeropuerto else "",
                                 "1" if valor.carretero else "",
                                 valor.tipoVehic.upper(),
                                 valor.lineaAutobus.upper(),
                                 valor.numeroEcono.upper(),
                                 valor.placas.upper(),
                                 "1" if valor.vehiculoAseg else "",
                                 "1" if valor.casaSeguridad else "",
                                 "1" if valor.centralAutobus else "",
                                 "1" if valor.ferrocarril else "",
                                 valor.empresa,
                                 "1" if valor.hotel else "",
                                 valor.nombreHotel,
                                 "1" if valor.puestosADispo else "",
                                 "1" if valor.juezCalif else "",
                                 "1" if valor.reclusorio else "",
                                 "1" if valor.policiaFede else "",
                                 "1" if valor.dif else "",
                                 "1" if valor.policiaEsta else "",
                                 "1" if valor.policiaMuni else "",
                                 "1" if valor.guardiaNaci else "",
                                 "1" if valor.fiscalia else "",
                                 "1" if valor.otrasAuto else "",
                                 "1" if valor.voluntarios else "",
                                 "1" if valor.otro else "",
                                 "1" if valor.presuntosDelincuentes else "",
                                 valor.numPresuntosDelincuentes if valor.numPresuntosDelincuentes != 0 else "",
                                 valor.municipio.upper(),
                                 valor.puntoEstra.upper(),
                                 valor.nacionalidad.upper(),
                                 valor.iso3,
                                 valor.nombre.upper(),
                                 valor.apellidos.upper(),
                                 valor.noIdentidad,
                                 valor.parentesco.upper(),
                                 valor.fechaNacimiento,
                                 "Hombre" if valor.sexo else "Mujer",
                                 "1" if valor.embarazo else "",
                                 valor.numFamilia if valor.numFamilia != 0 else "",
                                 valor.edad,
                                 ])
            
            worksheet.append(['Rescates Totales: ', valores.count()])

            response = HttpResponse(content = save_virtual_workbook(workbook), content_type='application/vnd.ms-excel.sheet.macroEnabled.12')
            response['Content-Disposition'] = 'attachment; filename={fecha}.xlsm'.format(fecha = fechaR)

            return response
    else:
        form = ExcelFormOr()
    return render(request, "generarExcel/generarArchivoExcelOrs.html", {"form" : form})

@csrf_exempt
def generarExcelORs(request):
    if(request.method == "POST"):
        form = ExcelFormOrs(request.POST)
        fechaR = request.POST["fechaDescarga"]

        if request.user.is_superuser:
            valores = RescatePunto.objects.filter(fecha= fechaR)
            oficinaR="Registros_ORs"
        else:
            oficinaR = request.POST["oficina"]
            # fechaR = datetime.datetime.strptime(f"{dia}/{mes}/{year}", "%d/%m/%Y").strftime('%d-%m-%y')
            # fechaR = datetime.datetime.strptime(f"{fechaFun}", "%d/%m/%Y").strftime('%d-%m-%y')

            valores = RescatePunto.objects.filter(fecha= fechaR).filter( oficinaRepre=oficinaR)
        
        
        workbook = opxl.load_workbook('tmp/dup.xlsm', read_only=False, keep_vba=True)
        worksheet = workbook.active
            
        for valor in valores:
            worksheet.append([valor.oficinaRepre, 
                                valor.fecha,
                                valor.hora,
                                valor.nombreAgente.upper(),
                                "1" if valor.aeropuerto else "",
                                "1" if valor.carretero else "",
                                valor.tipoVehic.upper(),
                                valor.lineaAutobus.upper(),
                                valor.numeroEcono.upper(),
                                valor.placas.upper(),
                                "1" if valor.vehiculoAseg else "",
                                "1" if valor.casaSeguridad else "",
                                "1" if valor.centralAutobus else "",
                                "1" if valor.ferrocarril else "",
                                valor.empresa,
                                "1" if valor.hotel else "",
                                valor.nombreHotel,
                                "1" if valor.puestosADispo else "",
                                "1" if valor.juezCalif else "",
                                "1" if valor.reclusorio else "",
                                "1" if valor.policiaFede else "",
                                "1" if valor.dif else "",
                                "1" if valor.policiaEsta else "",
                                "1" if valor.policiaMuni else "",
                                "1" if valor.guardiaNaci else "",
                                "1" if valor.fiscalia else "",
                                "1" if valor.otrasAuto else "",
                                "1" if valor.voluntarios else "",
                                "1" if valor.otro else "",
                                "1" if valor.presuntosDelincuentes else "",
                                valor.numPresuntosDelincuentes if valor.numPresuntosDelincuentes != 0 else "",
                                valor.municipio.upper(),
                                valor.puntoEstra.upper(),
                                valor.nacionalidad.upper(),
                                valor.iso3,
                                valor.nombre.upper(),
                                valor.apellidos.upper(),
                                valor.noIdentidad,
                                valor.parentesco.upper(),
                                valor.fechaNacimiento,
                                "Hombre" if valor.sexo else "Mujer",
                                "1" if valor.embarazo else "",
                                valor.numFamilia if valor.numFamilia != 0 else "",
                                valor.edad,
                                ])
        
        worksheet.append(['Rescates Totales: ', valores.count()])
        worksheet.append(['Oficina: ', oficinaR])

        response = HttpResponse(content = save_virtual_workbook(workbook), content_type='application/vnd.ms-excel.sheet.macroEnabled.12')
        response['Content-Disposition'] = 'attachment; filename={oficina}_{fecha}.xlsm'.format(fecha = fechaR, oficina=oficinaR)

        return response
    
    else:
        return render(request, "base/error404.html")






@login_required
@csrf_exempt
def generarExcelEdoFuerza(request):
    if(request.method == "POST"):

        if request.user.is_superuser:
                    valores = EstadoFuerza.objects.all().order_by('oficinaR')
                    
                    
                    workbook = opxl.load_workbook('tmp/eF.xlsm', read_only=False, keep_vba=True)
                    worksheet = workbook.active
                        
                    for valor in valores:
                        worksheet.append([valor.oficinaR, 
                                            valor.numPunto,
                                            valor.nomPuntoRevision,
                                            valor.tipoP,
                                            valor.ubicacion,
                                            valor.coordenadasTexto,
                                            valor.latitud,
                                            valor.longitud,
                                            valor.personalINM,
                                            valor.personalSEDENA,
                                            valor.personalMarina,
                                            valor.personalGuardiaN,
                                            valor.personalOTROS,
                                            valor.vehiculos,
                                            valor.seccion
                                            ])
                

                    response = HttpResponse(content = save_virtual_workbook(workbook), content_type='application/vnd.ms-excel.sheet.macroEnabled.12')
                    response['Content-Disposition'] = 'attachment; filename=EdoFuerza.xlsm'

                    return response
        else:

            print("no")
            
    
    else:
        return render(request, "base/error404.html")




@login_required

@csrf_exempt
def generarExcelUsuarios(request):
    if(request.method == "POST"):
        if request.user.is_superuser:
            valores = Usuario.objects.all().order_by("estado")
            workbook = opxl.load_workbook('tmp/us.xlsm', read_only=False, keep_vba=True)
            worksheet = workbook.active
            # print(valores)
            for valor in valores:
                estado = ""
                if valor.estado == '1':
                    estado = "AGUASCALIENTES"
                elif valor.estado == '2':
                    estado = "BAJA CALIFORNIA"
                elif valor.estado == '3':
                    estado = "BAJA CALIFORNIA SUR"
                elif valor.estado == '4':
                    estado = "CAMPECHE"
                elif valor.estado == '5':
                    estado = "COAHUILA"
                elif valor.estado == '6':
                    estado = "COLIMA"
                elif valor.estado == '7':
                    estado = "CHIAPAS"
                elif valor.estado == '8':
                    estado = "CHIHUAHUA"
                elif valor.estado == '9':
                    estado = "CDMX"
                elif valor.estado == '10':
                    estado = "DURANGO"
                elif valor.estado == '11':
                    estado = "GUANAJUATO"
                elif valor.estado == '12':
                    estado = "GUERRERO"
                elif valor.estado == '13':
                    estado = "HIDALGO"
                elif valor.estado == '14':
                    estado = "JALISCO"
                elif valor.estado == '15':
                    estado = "EDOMEX"
                elif valor.estado == '16':
                    estado = "MICHOACÁN"
                elif valor.estado == '17':
                    estado = "MORELOS"
                elif valor.estado == '18':
                    estado = "NAYARIT"
                elif valor.estado == '19':
                    estado = "NUEVO LEÓN"
                elif valor.estado == '20':
                    estado = "OAXACA"
                elif valor.estado == '21':
                    estado = "PUEBLA"
                elif valor.estado == '22':
                    estado = "QUERÉTARO"
                elif valor.estado == '23':
                    estado = "QUINTANA ROO"
                elif valor.estado == '24':
                    estado = "SAN LUIS POTOSÍ"
                elif valor.estado == '25':
                    estado = "SINALOA"
                elif valor.estado == '26':
                    estado = "SONORA"
                elif valor.estado == '27':
                    estado = "TABASCO"
                elif valor.estado == '28':
                    estado = "TAMAULIPAS"
                elif valor.estado == '29':
                    estado = "TLAXCALA"
                elif valor.estado == '30':
                    estado = "VERACRUZ"
                elif valor.estado == '31':
                    estado = "YUCATÁN"
                elif valor.estado == '32':
                    estado = "ZACATECAS"

                worksheet.append([
                    estado,
                    valor.nickname,
                    valor.nombre,
                    valor.apellido,
                ])


            response = HttpResponse(content = save_virtual_workbook(workbook), content_type='application/vnd.ms-excel.sheet.macroEnabled.12')
            response['Content-Disposition'] = 'attachment; filename=Usuarios.xlsm'
        else:
            
            print("no")
        return response
    
    else:
        return render(request, "base/error404.html")




@login_required

@csrf_exempt
def generarExcelPuntosI(request):
    if(request.method == "POST"):
        if request.user.is_superuser:
            valores = PuntosInternacion.objects.all().order_by("estadoPunto")
            workbook = opxl.load_workbook('tmp/pI.xlsm', read_only=False, keep_vba=True)
            worksheet = workbook.active
            # print(valores)
            for valor in valores:
                worksheet.append([
                    valor.estadoPunto,
                    valor.nombrePunto,
                    valor.tipoPunto,
                ])


            response = HttpResponse(content = save_virtual_workbook(workbook), content_type='application/vnd.ms-excel.sheet.macroEnabled.12')
            response['Content-Disposition'] = 'attachment; filename=PuntosInternacion.xlsm'
        else:
            
            print("no")
        return response
    
    else:
        return render(request, "base/error404.html")









@csrf_exempt    
def generarExcelFechas(request):
    if(request.method == "POST"):
        # print(request.POST)
        fechaI = request.POST["fechaInicio"]
        
        fechaF = request.POST["fechaFin"]

        fechaIN = datetime.strptime(f"{fechaI}", "%Y-%m-%d")
        fechaFN = datetime.strptime(f"{fechaF}", "%Y-%m-%d")

        # fecha_inicio = datetime(2024, 1, 1)
        # print(fecha_inicio)
        print(fechaIN)

        # fecha_inicio = datetime(yearI, mesI, diaI)
        # fecha_fin = datetime(diaF, mesF, yearF)

        array_fechas = [(fechaIN + timedelta(days=d)).strftime("%d-%m-%y") for d in range((fechaFN - fechaIN).days + 1)]

        # array_fechas = [(fechaIN + timedelta(days=d)).strftime("%d-%m-%y") for d in range((fechaIN - fechaFN).days + 1)]

        print(array_fechas)

        valores = RescatePunto.objects.filter(fecha__in= array_fechas)

        # if request.user.is_superuser:
        #     valores = valores = RescatePunto.objects.filter(fecha= fechaR)
        # else:
        #     oficinaR = request.POST["oficina"]
        #     # fechaR = datetime.datetime.strptime(f"{dia}/{mes}/{year}", "%d/%m/%Y").strftime('%d-%m-%y')
        #     # fechaR = datetime.datetime.strptime(f"{fechaFun}", "%d/%m/%Y").strftime('%d-%m-%y')
        
        workbook = opxl.load_workbook('tmp/dup.xlsm', read_only=False, keep_vba=True)
        worksheet = workbook.active
            
        for valor in valores:
            worksheet.append([valor.oficinaRepre, 
                                valor.fecha,
                                valor.hora,
                                valor.nombreAgente.upper(),
                                "1" if valor.aeropuerto else "",
                                "1" if valor.carretero else "",
                                valor.tipoVehic.upper(),
                                valor.lineaAutobus.upper(),
                                valor.numeroEcono.upper(),
                                valor.placas.upper(),
                                "1" if valor.vehiculoAseg else "",
                                "1" if valor.casaSeguridad else "",
                                "1" if valor.centralAutobus else "",
                                "1" if valor.ferrocarril else "",
                                valor.empresa,
                                "1" if valor.hotel else "",
                                valor.nombreHotel,
                                "1" if valor.puestosADispo else "",
                                "1" if valor.juezCalif else "",
                                "1" if valor.reclusorio else "",
                                "1" if valor.policiaFede else "",
                                "1" if valor.dif else "",
                                "1" if valor.policiaEsta else "",
                                "1" if valor.policiaMuni else "",
                                "1" if valor.guardiaNaci else "",
                                "1" if valor.fiscalia else "",
                                "1" if valor.otrasAuto else "",
                                "1" if valor.voluntarios else "",
                                "1" if valor.otro else "",
                                "1" if valor.presuntosDelincuentes else "",
                                valor.numPresuntosDelincuentes if valor.numPresuntosDelincuentes != 0 else "",
                                valor.municipio.upper(),
                                valor.puntoEstra.upper(),
                                valor.nacionalidad.upper(),
                                valor.iso3,
                                valor.nombre.upper(),
                                valor.apellidos.upper(),
                                valor.noIdentidad,
                                valor.parentesco.upper(),
                                valor.fechaNacimiento,
                                "Hombre" if valor.sexo else "Mujer",
                                "1" if valor.embarazo else "",
                                valor.numFamilia if valor.numFamilia != 0 else "",
                                valor.edad,
                                ])
        
        worksheet.append(['Rescates Totales: ', valores.count()])

        response = HttpResponse(content = save_virtual_workbook(workbook), content_type='application/vnd.ms-excel.sheet.macroEnabled.12')
        response['Content-Disposition'] = 'attachment; filename= "Rescates_{fecha1}-{fecha2}.xlsm"'.format(fecha1 = fechaI, fecha2=fechaF)

        return response
        # return render(request, "base/error404.html")
    
    else:
        return render(request, "base/error404.html")
    



@csrf_exempt
def politica_privacidad(request):
    
    return render(request, "politica/politica_privacidad.html")




@csrf_exempt    
def generarExcelFechasOR(request):
    if(request.method == "POST"):
        # print(request.POST)
        fechaI = request.POST["fechaInicio"]
        
        fechaF = request.POST["fechaFin"]

        fechaIN = datetime.strptime(f"{fechaI}", "%Y-%m-%d")
        fechaFN = datetime.strptime(f"{fechaF}", "%Y-%m-%d")
        
        # fecha_inicio = datetime(2024, 1, 1)
        # print(fecha_inicio)
        print(fechaIN)

        # fecha_inicio = datetime(yearI, mesI, diaI)
        # fecha_fin = datetime(diaF, mesF, yearF)

        array_fechas = [(fechaIN + timedelta(days=d)).strftime("%d-%m-%y") for d in range((fechaFN - fechaIN).days + 1)]

        # array_fechas = [(fechaIN + timedelta(days=d)).strftime("%d-%m-%y") for d in range((fechaIN - fechaFN).days + 1)]

        # print(array_fechas)
        oficinaR = request.POST["oficinaR"]
        #print(oficinaR)
        valores = RescatePunto.objects.filter(fecha__in= array_fechas, oficinaRepre= oficinaR )

        # if request.user.is_superuser:
        #     valores = valores = RescatePunto.objects.filter(fecha= fechaR)
        # else:
        #     oficinaR = request.POST["oficina"]
        #     # fechaR = datetime.datetime.strptime(f"{dia}/{mes}/{year}", "%d/%m/%Y").strftime('%d-%m-%y')
        #     # fechaR = datetime.datetime.strptime(f"{fechaFun}", "%d/%m/%Y").strftime('%d-%m-%y')
        
        workbook = opxl.load_workbook('tmp/dup.xlsm', read_only=False, keep_vba=True)
        worksheet = workbook.active
            
        for valor in valores:
            worksheet.append([valor.oficinaRepre, 
                                valor.fecha,
                                valor.hora,
                                valor.nombreAgente.upper(),
                                "1" if valor.aeropuerto else "",
                                "1" if valor.carretero else "",
                                valor.tipoVehic.upper(),
                                valor.lineaAutobus.upper(),
                                valor.numeroEcono.upper(),
                                valor.placas.upper(),
                                "1" if valor.vehiculoAseg else "",
                                "1" if valor.casaSeguridad else "",
                                "1" if valor.centralAutobus else "",
                                "1" if valor.ferrocarril else "",
                                valor.empresa,
                                "1" if valor.hotel else "",
                                valor.nombreHotel,
                                "1" if valor.puestosADispo else "",
                                "1" if valor.juezCalif else "",
                                "1" if valor.reclusorio else "",
                                "1" if valor.policiaFede else "",
                                "1" if valor.dif else "",
                                "1" if valor.policiaEsta else "",
                                "1" if valor.policiaMuni else "",
                                "1" if valor.guardiaNaci else "",
                                "1" if valor.fiscalia else "",
                                "1" if valor.otrasAuto else "",
                                "1" if valor.voluntarios else "",
                                "1" if valor.otro else "",
                                "1" if valor.presuntosDelincuentes else "",
                                valor.numPresuntosDelincuentes if valor.numPresuntosDelincuentes != 0 else "",
                                valor.municipio.upper(),
                                valor.puntoEstra.upper(),
                                valor.nacionalidad.upper(),
                                valor.iso3,
                                valor.nombre.upper(),
                                valor.apellidos.upper(),
                                valor.noIdentidad,
                                valor.parentesco.upper(),
                                valor.fechaNacimiento,
                                "Hombre" if valor.sexo else "Mujer",
                                "1" if valor.embarazo else "",
                                valor.numFamilia if valor.numFamilia != 0 else "",
                                valor.edad,
                                ])
        
        worksheet.append(['Rescates Totales: ', valores.count()])

        response = HttpResponse(content = save_virtual_workbook(workbook), content_type='application/vnd.ms-excel.sheet.macroEnabled.12')
        response['Content-Disposition'] = 'attachment; filename= "Rescates_{fecha1}-{fecha2}.xlsm"'.format(fecha1 = fechaI, fecha2=fechaF)

        return response
        # return render(request, "base/error404.html")
    
    else:
        return render(request, "base/error404.html")